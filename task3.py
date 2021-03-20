# Task №3

import openpyxl

class CreateExcelFile:

    def __init__(self, filename):
        self.filename_obj = openpyxl.load_workbook(filename)

    def __enter__(self):
        return self.filename_obj

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.filename_obj.close()

with CreateExcelFile("task_3.xlsx") as task3:
    task3_1 = task3.create_sheet("A", 0)
    task3_2 = task3.create_sheet("B")
    a = task3_1.cell(1, 1, "Cursor")
    b = task3_2.cell(2, 2, "Is the best")
    task3.save("task_3.xlsx")
    print(task3.sheetnames)
    print(a.value, b.value)